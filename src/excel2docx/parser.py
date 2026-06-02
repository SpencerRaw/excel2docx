"""Config-driven Excel parser.

Reads structured spreadsheets according to a schema config (YAML/JSON).
No domain knowledge baked in — the config defines:
  - Which sheets to read
  - Column mappings (name, type, aliases)
  - Aggregation rules (count by group, sum, etc.)
  - Row filters (skip empty, skip pattern)
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import yaml

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


def load_config(path: str | Path) -> dict:
    """Load parser config from YAML or JSON."""
    path = Path(path)
    if path.suffix in (".yaml", ".yml"):
        return yaml.safe_load(path.read_text())
    return json.loads(path.read_text())


def parse_excel(
    excel_path: str | Path,
    config: dict | str | Path,
) -> dict[str, Any]:
    """Parse an Excel workbook according to a config.

    Config structure:
        sheets:
          <sheet_name>:
            header_row: <int>        # row number (1-indexed) with headers
            data_start_row: <int>    # first data row
            data_end_row: <int|None> # last data row, None = scan until empty
            columns:
              <col_letter_or_index>:
                name: <str>          # output field name
                type: <str>          # int, float, str, date
                required: <bool>
                aliases: [<str>]     # alternate header names
            aggregates:
              group_by: <str>        # field to group by
              metric: count          # count | sum(<field>) | avg(<field>)

    Returns:
        {
            "sheets": {
                "<sheet_name>": {
                    "rows": [dict, ...],
                    "aggregates": {group: count, ...}
                }
            },
            "metadata": {"total_rows": N, "sheets_parsed": [...]}
        }
    """
    if isinstance(config, (str, Path)):
        config = load_config(config)

    if openpyxl is None:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    result: dict[str, Any] = {"sheets": {}, "metadata": {"total_rows": 0, "sheets_parsed": []}}

    for sheet_cfg in config.get("sheets", []):
        sheet_name = sheet_cfg["sheet_name"]
        if sheet_name not in wb.sheetnames:
            # Try fuzzy match
            matches = [s for s in wb.sheetnames if sheet_name.lower() in s.lower()]
            if matches:
                sheet_name = matches[0]
            else:
                continue

        ws = wb[sheet_name]
        rows = _parse_sheet(ws, sheet_cfg)
        aggregates = _compute_aggregates(rows, sheet_cfg.get("aggregates", {}))

        result["sheets"][sheet_cfg.get("output_name", sheet_name)] = {
            "rows": rows,
            "aggregates": aggregates,
        }
        result["metadata"]["total_rows"] += len(rows)
        result["metadata"]["sheets_parsed"].append(sheet_cfg.get("output_name", sheet_name))

    return result


def _parse_sheet(ws, cfg: dict) -> list[dict]:
    """Parse a single sheet."""
    header_row = cfg.get("header_row", 1)
    data_start = cfg.get("data_start_row", header_row + 1)
    data_end = cfg.get("data_end_row")  # None = scan until empty
    columns_cfg = cfg.get("columns", [])
    skip_empty_rows = cfg.get("skip_empty_rows", True)
    skip_pattern = cfg.get("skip_pattern")  # regex to skip rows

    # Build header → output name mapping
    header_map = {}
    header_values = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]

    for col_cfg in columns_cfg:
        output_name = col_cfg.get("name", col_cfg.get("col", ""))
        aliases = col_cfg.get("aliases", [output_name])

        for alias in aliases:
            for idx, h in enumerate(header_values):
                if h and alias.lower() == h.lower():
                    header_map[idx] = {
                        "name": output_name,
                        "type": col_cfg.get("type", "str"),
                        "required": col_cfg.get("required", False),
                    }
                    break

    # Parse data rows
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True)):
        if data_end and i + data_start > data_end:
            break

        # Skip completely empty rows
        if skip_empty_rows and all(v is None for v in row):
            continue

        entry = {}
        for col_idx, hdr_info in header_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None and hdr_info["required"]:
                entry[hdr_info["name"]] = ""
            elif val is not None:
                entry[hdr_info["name"]] = _cast_value(val, hdr_info["type"])
            else:
                entry[hdr_info["name"]] = ""

        # Skip pattern check
        if skip_pattern:
            import re
            if any(re.search(skip_pattern, str(v)) for v in entry.values()):
                continue

        rows.append(entry)

    return rows


def _cast_value(val: Any, dtype: str) -> Any:
    """Cast a cell value to the target type."""
    if dtype == "str":
        return str(val).strip()
    if dtype == "int":
        try:
            return int(float(str(val).replace(",", "")))
        except (ValueError, TypeError):
            return 0
    if dtype == "float":
        try:
            return float(str(val).replace(",", "").replace("$", ""))
        except (ValueError, TypeError):
            return 0.0
    if dtype == "date":
        from datetime import datetime
        if isinstance(val, datetime):
            return val
        return str(val)
    return str(val)


def _compute_aggregates(rows: list[dict], agg_cfg: dict) -> dict:
    """Compute aggregations over parsed rows."""
    if not agg_cfg:
        return {}

    group_by = agg_cfg.get("group_by", "")
    metric = agg_cfg.get("metric", "count")

    if not group_by:
        return {}

    groups: dict[str, Any] = {}
    for row in rows:
        key = str(row.get(group_by, "Unknown")).strip()
        if key not in groups:
            groups[key] = [] if metric.startswith(("sum(", "avg(")) else 0

        if metric == "count":
            groups[key] += 1
        elif metric.startswith("sum("):
            field = metric[4:-1]
            groups[key].append(row.get(field, 0))
        elif metric.startswith("avg("):
            field = metric[4:-1]
            groups[key].append(row.get(field, 0))

    # Finalize
    if metric.startswith("sum("):
        for k in groups:
            groups[k] = sum(v for v in groups[k] if isinstance(v, (int, float)))
    elif metric.startswith("avg("):
        for k in groups:
            vals = [v for v in groups[k] if isinstance(v, (int, float))]
            groups[k] = sum(vals) / len(vals) if vals else 0

    return groups
