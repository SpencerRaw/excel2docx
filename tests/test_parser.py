"""Tests for excel2docx parser."""

import tempfile
from pathlib import Path

import pytest

from excel2docx.parser import parse_excel, load_config


def test_load_config_yaml():
    """Config can be loaded from YAML."""
    cfg = load_config("examples/business_daily_report.yaml")
    assert "parser" in cfg
    assert "template" in cfg
    assert cfg["transform"]["mode"] == "rules"


def test_parse_excel_minimal():
    """Parse a minimal Excel file with known structure."""
    # Create a test Excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Log"
    ws.append(["Ref ID", "Category", "Department", "Description"])
    ws.append(["R001", "Incident", "Security", "Door alarm triggered"])
    ws.append(["R002", "Report", "Operations", "Shift change complete"])
    ws.append(["R003", "Incident", "Security", "Visitor badge issued"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        tmp_path = f.name

    config = {
        "sheets": [{
            "sheet_name": "Daily Log",
            "output_name": "logs",
            "header_row": 1,
            "data_start_row": 2,
            "columns": [
                {"col": 0, "name": "ref_id", "type": "str"},
                {"col": 1, "name": "category", "type": "str"},
                {"col": 2, "name": "department", "type": "str"},
                {"col": 3, "name": "description", "type": "str"},
            ],
            "aggregates": {"group_by": "department", "metric": "count"},
        }],
    }

    result = parse_excel(tmp_path, config)
    rows = result["sheets"]["logs"]["rows"]

    assert len(rows) == 3
    assert rows[0]["ref_id"] == "R001"
    assert rows[1]["category"] == "Report"
    assert result["metadata"]["total_rows"] == 3
    assert result["sheets"]["logs"]["aggregates"]["Security"] == 2

    Path(tmp_path).unlink()


def test_parse_with_fuzzy_sheet_match():
    """Fuzzy sheet name matching works."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data 2024"
    ws.append(["Product", "Revenue"])
    ws.append(["Widget", 100])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        tmp_path = f.name

    config = {
        "sheets": [{
            "sheet_name": "Sales Data",  # partial match
            "header_row": 1,
            "data_start_row": 2,
            "columns": [
                {"col": 0, "name": "product", "type": "str"},
                {"col": 1, "name": "revenue", "type": "float"},
            ],
        }],
    }

    result = parse_excel(tmp_path, config)
    assert result["metadata"]["total_rows"] == 1

    Path(tmp_path).unlink()


def test_type_casting():
    """Float and int types are properly cast."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Count", "Amount"])
    ws.append(["Item A", 5, "1,234.56"])
    ws.append(["Item B", 10, "$500"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        tmp_path = f.name

    config = {
        "sheets": [{
            "sheet_name": "Data",
            "header_row": 1,
            "data_start_row": 2,
            "columns": [
                {"col": 0, "name": "name", "type": "str"},
                {"col": 1, "name": "count", "type": "int"},
                {"col": 2, "name": "amount", "type": "float"},
            ],
        }],
    }

    result = parse_excel(tmp_path, config)
    rows = result["sheets"]["Data"]["rows"]
    assert rows[0]["count"] == 5
    assert rows[0]["amount"] == 1234.56
    assert rows[1]["amount"] == 500.0

    Path(tmp_path).unlink()
